import openpyxl as xl
import random


classes_wb = xl.load_workbook('classes.xlsx')
classes_sheet = classes_wb['Sheet1']

# unused
	# class_name = '' 
	# class_hours = 0
	# class_names = 1
	# classes = ''

days = ("Monday", "Tuesday", "Wednesday", "Thusday", "Friday", "Saturday")
hours = (1, 2, 3, 4, 5)
loop_rows = classes_sheet.max_row + 1
loop_columns = classes_sheet.max_column + 1
temp_row_worksheet_to_array = []
worksheet_to_array = []
temp_time_table = []
time_table = []
professor_hours = 0
class_lessons = []
class_lessons_array = []
day_time_table = []
week_time_table = []
class_time_table = []


for row in range(2, loop_rows):
	for column in range(1, loop_columns):
		cell_value = classes_sheet.cell(row, column).value
		temp_row_worksheet_to_array.append(cell_value)
	worksheet_to_array.append(temp_row_worksheet_to_array)
	temp_row_worksheet_to_array = []

for row in worksheet_to_array:
	row_lenght = len(row)
	row_number = worksheet_to_array.index(row)
	for cell_index_p1 in range(2, row_lenght + 1):
		cell_index = cell_index_p1 - 1
		cell = row[cell_index]
		if cell_index != 0:
			loop_statement = cell_index % 2
			if loop_statement == 0:
				professor_hours += int(cell)
			else:
				professor = cell
			for hour in range(1, professor_hours+1):
				class_lessons.append(professor)
			professor_hours = 0
	class_lessons_array.append(class_lessons)
	class_lessons = []


for class_lessons in range(0, len(class_lessons_array)):
	lessons_left = (class_lessons_array[class_lessons])
	for day in days:
		for lesson in hours:
			random_lesson = random.choice(lessons_left)
			lessons_left.remove(random_lesson)
			# print(random_lesson)

			day_time_table.append(random_lesson)
		day_time_table.append(day)
		class_time_table.append(day_time_table)
		day_time_table = []
	time_table.append(class_time_table)
	class_time_table = []

print(time_table)

# for lesson in time_table[0]:
# 	index = time_table.index(lesson)
# 	for row in time_table:
# 		if

# Debug
# print(class_lessons_array)